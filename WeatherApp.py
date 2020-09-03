from tkinter import *
import tkinter.messagebox
import datetime
running = False
import requests
from openpyxl import Workbook
from time import sleep
now = datetime.datetime.now()
import webbrowser
running = True  # Global flag
count = 0
list = []
time_list = []
city_list=[]
unit_list=[]

def scanning():
    city_name = city_f.get()
    unit_is = unit.get()
    global list
    global time_list
    global count
    if running:  # Only do this if the Stop button has not been clicked
        if unit_is == 'C':
            delay = int(second.get())
            response = requests.get(
                'http://api.weatherstack.com/current?access_key=d6734e2398411020f27ff8c6cdfbf374& query= ' + city_name)
            json_object = response.json()
            temp_k = json_object['current']['temperature']
            list.append(temp_k)
            time = now.strftime("%Y-%m-%d %H:%M:%S")
            time_list.append(time)
            city_list.append(city_name)
            unit_list.append(unit_is)
            count = count + 1
            sleep(delay)
        elif unit_is == 'F':
            delay = int(second.get())
            response = requests.get(
                'http://api.weatherstack.com/current?access_key=d6734e2398411020f27ff8c6cdfbf374& query= ' + city_name)
            json_object = response.json()
            temp_k = json_object['current']['temperature']
            temp_f = (temp_k * 1.8) + 32
            list.append(temp_f)
            city_list.append(city_name)
            unit_list.append(unit_is)
            time = now.strftime("%Y-%m-%d %H:%M:%S")
            time_list.append(time)
            count = count + 1
            sleep(delay)
        root.after(1000, scanning)

def stop():
    """Stop scanning by setting the global flag to False."""
    global running
    running = False
    book = Workbook()
    sheet = book.active
    sheet['A1'] = 'City Name'
    sheet['B1'] = 'Temp'
    sheet['C1'] = 'Time'
    sheet['D1'] = 'Unit'
    row = 2
    column = 1
    temp_row = 2
    temp_column = 2
    time_row = 2
    time_column = 3
    unit_row = 2
    unit_column = 4
    global list
    global time_list
    for i in range(len(city_list)):
        sheet.cell(row=row, column=column).value = city_list[i]
        row = row + 1

    for i in range(len(list)):
        sheet.cell(row=temp_row, column=temp_column).value = list[i]
        temp_row = temp_row + 1
    for i in range(len(time_list)):
        sheet.cell(row=time_row, column=time_column).value = time_list[i]
        time_row = time_row + 1
    unit_is = unit.get()
    for i in range(len(unit_list)):
        if unit_is == 'C':
            sheet.cell(row=unit_row, column=unit_column).value = unit_list[i]
            unit_row = unit_row + 1
        elif unit_is == 'F':
            sheet.cell(row=unit_row, column=unit_column).value = unit_list[i]
            unit_row = unit_row + 1
    book.save('Weather_Report.xlsx')
    tkinter.messagebox.showinfo("Successful", "Process has been stop! Report is Generated in Weather_Report.xlsx")

def start():
    """Enable scanning by setting the global flag to True."""
    global running
    running = True
    if len(city_f.get())==0 or len(unit.get())==0 or len(second.get())==0:
        tkinter.messagebox.showinfo("Error", "Entry is Empty")
    
    else:
        scanning()
        tkinter.messagebox.showinfo("Successful", "Process has been Started!")

def callback(url):
    webbrowser.open_new(url)


root = Tk()
root.title("WeatherApp By Samir")
root.configure(background="#a1dbcd")
root.geometry("500x480")
label = Label(root, text="Weather Script By Samir", fg='#a1dbcd', bg='#383a39')
label1 = Label(root, text="Enter the City :", fg='black', bg='#a1dbcd')
label3 = Label(root, text="Type 'C' for Degree or 'F' for Fahrenheit  :", fg='blue', bg='#a1dbcd')
label4 = Label(root, text="Delay in seconds :", fg='black', bg='#a1dbcd')
city_f = Entry(root)
unit = Entry(root)
second = Entry(root)
b1 = Button(root, text="Tell Weather!", bg='#383a39', fg='#a1dbcd', command=start)
b2 = Button(root, text="Stop", bg='#383a39', fg='#a1dbcd', command=stop)
b3 = Button(root, text="Exit", bg='#383a39', fg='#a1dbcd', command=root.destroy)
label.grid(row=0, column=2)
label1.grid(row=2, column=2)
label3.grid(row=5, column=2)
label4.grid(row=8, column=2)
city_f.grid(row=4, column=2, ipadx="180")
unit.grid(row=7, column=2, ipadx="20")
second.grid(row=10, column=2, ipadx="1")
b1.grid(row=11, column=2, )
b2.grid(row=13, column=2)
b3.grid(row=14, column=2)

link1 = Label(root, text="GitHub", fg="blue", cursor="hand2")
link1.grid(row=15, column=2)
link1.bind("<Button-1>", lambda e: callback("https://github.com/samir321-pixel"))
root.mainloop()
