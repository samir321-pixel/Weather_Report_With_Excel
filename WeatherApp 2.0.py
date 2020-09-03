import openpyxl
import xlrd
import requests
from time import sleep, time

wb = openpyxl.load_workbook('C:\\Users\\Criminal\\PycharmProjects\\WeatherApp-master\\Weather.xlsx')
ws = wb['Sheet1']
book = xlrd.open_workbook('C:\\Users\\Criminal\\PycharmProjects\\WeatherApp-master\\Weather.xlsx')
max_nb_row = 0
for sheet in book.sheets():
    max_nb_row = max(max_nb_row, sheet.nrows)
max = max_nb_row - 1
#print('Number Of Filled Rows', max)
temp_row = 2
temp_column = 2
row = 2
column = 1
update_row = 2
update_column = 4
unit_row = 2
unit_column = 3
x = 1

def run():
    print('Process Has Been Started................')
    global update_row
    global row
    global unit_row
    global temp_row
    go = True
    while go == True:
        for i in range(max):
            if ws.cell(row=update_row, column=update_column).value == 1:
                if ws.cell(row=unit_row, column=unit_column).value == 'C':
                    response = requests.get(
                        'http://api.weatherstack.com/current?access_key=d6734e2398411020f27ff8c6cdfbf374& query= ' + ws.cell(
                            row=row, column=column).value)
                    json_object = response.json()
                    temp_k = json_object['current']['temperature']
                    #print('city : ' + ws.cell(row=row, column=column).value + ' has Temperature Requested in: ' + ws.cell(row=unit_row, column=unit_column).value + ' is ' + str(temp_k))
                    row = row + 1
                    update_row = update_row + 1
                    unit_row = unit_row + 1
                    ws.cell(row=temp_row, column=temp_column).value = temp_k
                    wb.save('C:\\Users\\Criminal\\PycharmProjects\\WeatherApp-master\\Weather.xlsx')
                    #print("Excel Updated")
                    temp_row = temp_row + 1
                elif ws.cell(row=unit_row, column=unit_column).value == 'F':
                    response = requests.get(
                        'http://api.weatherstack.com/current?access_key=d6734e2398411020f27ff8c6cdfbf374& query= ' + ws.cell(
                            row=row, column=column).value)
                    json_object = response.json()
                    temp_k = json_object['current']['temperature']
                    temp_f = (temp_k * 1.8) + 32
                    #print('city : ' + ws.cell(row=row, column=column).value + ' has Temperature Requested in: ' + ws.cell(row=unit_row, column=unit_column).value + ' is ' + str(temp_f))
                    row = row + 1
                    update_row = update_row + 1
                    unit_row = unit_row + 1
                    ws.cell(row=temp_row, column=temp_column).value = temp_f
                    wb.save('C:\\Users\\Criminal\\PycharmProjects\\WeatherApp-master\\Weather.xlsx')
                    temp_row = temp_row + 1
            elif ws.cell(row=update_row, column=update_column).value == 0:
                #print('We will not update the city because Not Allowed : ', ws.cell(row=row, column=column).value)
                constant = ws.cell(row=temp_row, column=temp_column).value
                #print('For City: ' + ws.cell(row=row,column=column).value + ' Keeping Temperature same because Not Allowed : ' + str(constant))
                ws.cell(row=temp_row, column=temp_column).value = constant
                wb.save('C:\\Users\\Criminal\\PycharmProjects\\WeatherApp-master\\Weather.xlsx')
                temp_row = temp_row + 1
                row = row + 1
                update_row = update_row + 1
                unit_row = unit_row + 1
        sleep(1)
        unit_row = 2
        update_row = 2
        row = 2
        temp_row = 2

run()
